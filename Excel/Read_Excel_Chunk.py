import pandas as pd
from openpyxl import load_workbook

'''
    pd.read_excel()은 더 이상 chunksize 를 지원하지 않는다.
    따라서 특정 row수만큼 반복하여 끊어 읽기 위해서는 loop을 사용해야 한다.
    This code shows how to whole read excel file chunk by chunk with pd.read_excel.
'''


def readList(import_file, chunk_count, chunk):
'''
이전에 읽은 rows는 건너뛰어 읽는 함수
Function for skipping old rows
'''
    print("...Skipped "+str(chunk_count*chunk)+" rows from first row")
    comp = pd.read_excel(import_path, header=0, sheet_name=0, skiprows=range(0,chunk_count*chunk), nrows=chunk)


# input 엑셀 파일 경로 
input_path = 'C:\\Users\\mypc\\Desktop\\input.xlsx'
# output 엑셀 파일 경로 (확장자 .xlsx는 생략할 것)
output_path = 'C:\\Users\\mypc\\Desktop\\output_'
# 끊어읽는 단위 
chunk = 100

# 전체 row 수, iteration 수 구하기
wb = load_workbook(input_path)
sheet = wb.worksheets[0]
row_count = sheet.max_row       # input 파일의 전체 row 수
if row_count%chunk == 0:
    no_iter = row_count//chunk  # 끊어 읽기 위한 iteration 횟수
else:
    no_iter = row_count//chunk + 1
del wb
del sheet

for chunk_count in range(0,no_iter):
    print("......Reading Iteration: "+str(chunk_count+1)+"/"+str(no_iter)+"......")
    cert, comp = readList(input_path, chunk_count, r)
